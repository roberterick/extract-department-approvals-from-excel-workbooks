import openpyxl as opx
import os,shutil
import locale
locale.setlocale(locale.LC_ALL, '')

EXTRACT_PATH=r'path here'
BG_PART='2022-2023 masterlist department approvals'

'''
given a directory of workbooks named using approver emails
scroll through the excel workbooks extracting on/off status
format resulting file for upload to Blackbaud Financial Edge
'''

class app(object):
    def __init__(self):
        self.datalist=[]
        shortlist=[x for x in os.listdir(EXTRACT_PATH)]
        shortlist=filter(lambda x:x.endswith('xlsx'),shortlist)
        shortlist=filter(lambda x:x.find('@')>-1,shortlist)
        shortlist.sort()
        longlist=[os.path.join(EXTRACT_PATH,x) for x in shortlist]
        for x in longlist:
            print x
            self.get_data(x)
        #
        self.datalist.sort()
        self.datalist.insert(0,('fund','at','ad'))
        self.export_data()

    def get_data(self,fpth):
        print 'getting data for %s'%fpth
        email=os.path.basename(fpth)[:-5]
        wb=opx.load_workbook(fpth)
        for sn in wb.sheetnames:
            ws=wb[sn]
            for row in ws.iter_rows(min_row=9,max_col=3,values_only=True):
                if hasattr(row[0],'__len__') and len(row[0])==4:
                    fund=row[0][:4]
                    yn=row[2][0]
                    tup=(fund,BG_PART,'approved by %s as %s'%(email,yn))
                    self.datalist+=[tup]

    def export_data(self):
        fpth=os.path.join(EXTRACT_PATH,'__result.xlsx')
        print fpth
        wb=opx.Workbook()
        ws=wb.active
        ws.name="upload"
        #
        r=0
        for t in self.datalist:
            for c in range(3):
                ws.cell(row=r+1,column=c+1,value=t[c])
            #
            r+=1
        #
        wb.save(fpth)




app()

